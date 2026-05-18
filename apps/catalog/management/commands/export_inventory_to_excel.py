"""
export_inventory_to_excel — vuelca el estado actual del inventario a un Excel.

Genera un archivo con una hoja por categoria principal (raiz). Cada fila es
una variante. Las columnas coinciden con las que espera sync_inventory_from_excel,
de modo que el Excel exportado sirve como entrada del importador (round-trip).

Uso:
    python manage.py export_inventory_to_excel inventario_actual.xlsx
    python manage.py export_inventory_to_excel /tmp/snapshot.xlsx --include-inactive
"""
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from apps.catalog.models import Category, ProductVariant


HEADERS = ['sku', 'product_slug', 'nombre', 'color', 'precio', 'stock', 'activo']


class Command(BaseCommand):
    help = 'Exporta el inventario (variantes) a un archivo Excel.'

    def add_arguments(self, parser):
        parser.add_argument('output', type=str, help='Ruta de salida (.xlsx)')
        parser.add_argument('--include-inactive', action='store_true',
                            help='Incluir variantes con is_active=False.')

    def handle(self, *args, **options):
        try:
            from openpyxl import Workbook
        except ImportError:
            raise CommandError('openpyxl no esta instalado. pip install openpyxl')

        out = Path(options['output'])
        include_inactive = options['include_inactive']

        wb = Workbook()
        wb.remove(wb.active)

        # Una hoja por categoria raiz
        roots = Category.objects.filter(parent__isnull=True).order_by('order', 'name')
        total_rows = 0

        # Caracteres prohibidos en nombres de hoja Excel: : \ / ? * [ ]
        _illegal = set(r':\/?*[]')

        def _sanitize(name):
            cleaned = ''.join(' ' if c in _illegal else c for c in name)
            return cleaned[:31].strip() or 'Hoja'

        for root in roots:
            sheet_name = _sanitize(root.name)
            ws = wb.create_sheet(title=sheet_name)
            ws.append(HEADERS)

            qs = ProductVariant.objects.filter(
                product__category__parent=root,
            ).select_related('product').order_by('product__slug', 'color')
            if not include_inactive:
                qs = qs.filter(is_active=True)

            rows_in_sheet = 0
            for v in qs:
                ws.append([
                    v.sku,
                    v.product.slug,
                    v.name,
                    v.color,
                    float(v.price),
                    v.stock,
                    'true' if v.is_active else 'false',
                ])
                rows_in_sheet += 1

            # Ancho de columna basico
            for col_idx, _ in enumerate(HEADERS, start=1):
                ws.column_dimensions[chr(64 + col_idx)].width = 18

            self.stdout.write(f'  hoja "{sheet_name}": {rows_in_sheet} variantes')
            total_rows += rows_in_sheet
            if rows_in_sheet == 0:
                # Dejar la hoja con solo el header — utilizable para llenar a mano.
                pass

        if not roots:
            ws = wb.create_sheet(title='Inventario')
            ws.append(HEADERS)
            self.stdout.write(self.style.WARNING('  no hay categorias raiz; hoja vacia "Inventario" creada.'))

        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        self.stdout.write(self.style.SUCCESS(
            f'\nOK: {total_rows} variantes exportadas a {out}'
        ))
