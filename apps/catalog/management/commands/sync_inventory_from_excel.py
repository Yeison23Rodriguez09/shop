"""
sync_inventory_from_excel — sincroniza inventario (precio/stock/variantes) desde Excel.

El Excel es la unica fuente de verdad para inventario. Una hoja por categoria
principal. Cada fila = una variante (color) de un producto.

Columnas minimas (header en fila 1, case-insensitive, espacios opcionales):
    sku           identificador unico de la variante
    product_slug  slug del Product (debe existir; sembrado por sync_content)
    nombre        etiqueta visible de la variante (opcional)
    precio        decimal
    stock         entero >= 0
    color         etiqueta del color (vacio = variante default)
    activo        true/1/si/yes para activo, false/0/no para inactivo

Reglas:
  - Idempotente: usa update_or_create por SKU.
  - Si product_slug no existe en BD: warning, fila omitida.
  - Si SKU duplicado dentro del mismo Excel: error, fila omitida.
  - Si (product, color) ya existe con OTRO sku: error (constraint unique_together).
  - Solo afecta variantes presentes en el Excel; variantes en BD que ya no
    aparecen NO se borran automaticamente (se reportan como huerfanas).
    Para desactivarlas, marca activo=False en su fila.

Uso:
    python manage.py sync_inventory_from_excel inventario.xlsx
    python manage.py sync_inventory_from_excel inventario.xlsx --dry-run
"""
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.catalog.models import Product, ProductVariant


REQUIRED_COLUMNS = ('sku', 'product_slug', 'precio', 'stock')
OPTIONAL_COLUMNS = ('nombre', 'color', 'activo')


def _norm(s):
    """Normaliza header: lowercase, sin espacios laterales, sin acentos basicos."""
    if s is None:
        return ''
    return str(s).strip().lower().replace('á', 'a').replace('é', 'e') \
                  .replace('í', 'i').replace('ó', 'o').replace('ú', 'u')


def _parse_bool(v, default=True):
    if v is None:
        return default
    s = str(v).strip().lower()
    if s in ('1', 'true', 'si', 'sí', 'yes', 'y', 'activo', 'verdadero'):
        return True
    if s in ('0', 'false', 'no', 'inactivo', 'falso'):
        return False
    return default


class Command(BaseCommand):
    help = 'Sincroniza inventario (precio/stock/variantes por color) desde un Excel.'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Ruta al .xlsx')
        parser.add_argument('--dry-run', action='store_true',
                            help='Muestra cambios sin escribir.')
        parser.add_argument('--verbose', action='store_true')

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError('openpyxl no esta instalado. pip install openpyxl')

        path = Path(options['excel_file'])
        if not path.is_file():
            raise CommandError(f'No existe el archivo: {path}')

        self.dry_run = options['dry_run']
        self.verbose = options['verbose']

        wb = load_workbook(path, data_only=True, read_only=True)
        self.stats = {
            'created': 0, 'updated': 0, 'unchanged': 0,
            'skipped_no_product': 0, 'skipped_dup_sku': 0,
            'errors': 0, 'orphans': 0,
        }

        seen_skus_global = set()
        all_variants_in_excel = []  # [(sheet, row_idx, parsed_dict), ...]

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self.stdout.write(self.style.HTTP_INFO(f'\n=== Hoja: {sheet_name} ==='))

            rows = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration:
                self.stdout.write(self.style.WARNING('  hoja vacia'))
                continue

            header = {_norm(h): idx for idx, h in enumerate(header_row) if h is not None}
            missing = [c for c in REQUIRED_COLUMNS if c not in header]
            if missing:
                self._err(f'  hoja "{sheet_name}": faltan columnas requeridas {missing} '
                          f'(detectadas: {list(header.keys())}) — hoja omitida.')
                continue

            for row_idx, row in enumerate(rows, start=2):
                if all(c is None for c in row):
                    continue
                parsed = self._parse_row(row, header, sheet_name, row_idx)
                if parsed is None:
                    continue
                # Detectar SKU duplicado entre filas
                sku = parsed['sku']
                if sku in seen_skus_global:
                    self._err(f'  [{sheet_name}!row {row_idx}] SKU duplicado: "{sku}" — fila omitida.')
                    self.stats['skipped_dup_sku'] += 1
                    continue
                seen_skus_global.add(sku)
                all_variants_in_excel.append((sheet_name, row_idx, parsed))

        # Aplicar todo en una transaccion (para que un error no deje BD a medias).
        if self.dry_run:
            for sheet, ridx, p in all_variants_in_excel:
                self.stdout.write(f'  [dry-run] {sheet}!row {ridx}: '
                                  f'{p["sku"]} -> {p["product_slug"]} '
                                  f'color={p["color"] or "—"} precio={p["price"]} stock={p["stock"]}')
        else:
            with transaction.atomic():
                for sheet, ridx, p in all_variants_in_excel:
                    self._apply_variant(p, sheet, ridx)

        self._report_orphans(seen_skus_global)
        self._print_summary()

    # ──────────────────────────────────────────────────────────
    def _parse_row(self, row, header, sheet, ridx):
        """Convierte una fila a dict normalizado. Retorna None si invalida."""
        def cell(name):
            i = header.get(name)
            return row[i] if i is not None and i < len(row) else None

        sku_raw = cell('sku')
        slug_raw = cell('product_slug')
        if sku_raw is None or str(sku_raw).strip() == '':
            return None
        if slug_raw is None or str(slug_raw).strip() == '':
            self._err(f'  [{sheet}!row {ridx}] product_slug vacio — fila omitida.')
            self.stats['errors'] += 1
            return None

        try:
            price = Decimal(str(cell('precio') or '0'))
        except (InvalidOperation, TypeError):
            self._err(f'  [{sheet}!row {ridx}] precio invalido: {cell("precio")!r}')
            self.stats['errors'] += 1
            return None

        stock_raw = cell('stock')
        try:
            stock = max(0, int(stock_raw))
        except (TypeError, ValueError):
            self._err(f'  [{sheet}!row {ridx}] stock invalido: {stock_raw!r}')
            self.stats['errors'] += 1
            return None

        return {
            'sku': str(sku_raw).strip(),
            'product_slug': str(slug_raw).strip(),
            'name': str(cell('nombre') or '').strip(),
            'color': str(cell('color') or '').strip(),
            'price': price,
            'stock': stock,
            'is_active': _parse_bool(cell('activo'), default=True),
        }

    def _apply_variant(self, p, sheet, ridx):
        try:
            product = Product.objects.get(slug=p['product_slug'])
        except Product.DoesNotExist:
            self.stdout.write(self.style.WARNING(
                f'  [{sheet}!row {ridx}] product_slug "{p["product_slug"]}" '
                f'no existe en BD — fila omitida.'
            ))
            self.stats['skipped_no_product'] += 1
            return

        defaults = {
            'product': product,
            'color': p['color'],
            'name': p['name'],
            'price': p['price'],
            'stock': p['stock'],
            'is_active': p['is_active'],
        }

        try:
            obj, created = ProductVariant.objects.update_or_create(
                sku=p['sku'], defaults=defaults,
            )
        except Exception as e:
            self._err(f'  [{sheet}!row {ridx}] error guardando variante {p["sku"]}: {e}')
            self.stats['errors'] += 1
            return

        if created:
            self.stats['created'] += 1
            self._ok(f'  [+] {p["sku"]} -> {product.slug} ({p["color"] or "default"})')
        else:
            self.stats['updated'] += 1
            if self.verbose:
                self._ok(f'  [~] {p["sku"]} actualizado')

    def _report_orphans(self, seen_skus):
        """Variantes en BD que no estan en el Excel — solo se reportan."""
        orphans = ProductVariant.objects.exclude(sku__in=seen_skus)
        n = orphans.count()
        if n:
            self.stats['orphans'] = n
            self.stdout.write(self.style.WARNING(
                f'\n  {n} variante(s) en BD NO estan en el Excel actual '
                f'(no se borran automaticamente; usa activo=False para desactivarlas):'
            ))
            for v in orphans[:10]:
                self.stdout.write(f'    - {v.sku} ({v.product.slug} / {v.color or "default"})')
            if n > 10:
                self.stdout.write(f'    ... y {n - 10} mas')

    def _print_summary(self):
        s = self.stats
        self.stdout.write('\n' + '=' * 60)
        self.stdout.write(f'  Variantes creadas:    {s["created"]}')
        self.stdout.write(f'  Variantes actualizadas: {s["updated"]}')
        self.stdout.write(f'  Sin product_slug en BD: {s["skipped_no_product"]}')
        self.stdout.write(f'  SKUs duplicados omitidos: {s["skipped_dup_sku"]}')
        if s['errors']:
            self.stdout.write(self.style.ERROR(f'  Errores: {s["errors"]}'))
        if s['orphans']:
            self.stdout.write(self.style.WARNING(f'  Huerfanos en BD: {s["orphans"]}'))
        self.stdout.write('=' * 60)
        if self.dry_run:
            self.stdout.write(self.style.WARNING('  ** DRY RUN — no se persistio nada **'))

    def _err(self, msg):
        self.stderr.write(self.style.ERROR(msg))

    def _ok(self, msg):
        self.stdout.write(self.style.SUCCESS(msg))
